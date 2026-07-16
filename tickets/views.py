from django.contrib.auth import authenticate, get_user_model, login as auth_login, logout as django_logout
from django.contrib.auth.decorators import login_required
import threading
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Avg, F, ExpressionWrapper, DurationField
from django.conf import settings
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.decorators import authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from datetime import timedelta, datetime, date
from django.core.mail import EmailMessage
from django.db.models.functions import TruncDate
from .models import RelacionUsuarioSistema
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl

import csv
import io
import json

from rest_framework.authentication import BaseAuthentication
from rest_framework.exceptions import AuthenticationFailed

from .models import (
    Usuario, Sistema, Modulo, Documento, Prioridad, Estado, Categoria,
    Ticket, ChatterEntry, TicketTimeLog, ConocimientoEntry, Token, Departamento
)

from .serializers import (
    UsuarioSerializer, UsuarioInputSerializer, UsuarioUpdateSerializer,
    SistemaSerializer, ModuloSerializer, DocumentoSerializer,
    PrioridadSerializer, EstadoSerializer, CategoriaSerializer,
    TicketSerializer, TicketInputSerializer, TicketUpdateSerializer,
    ChatterEntrySerializer, ChatterInputSerializer,
    TimeLogSerializer, ConocimientoSerializer,
)
from . import resend_email


# ─────────────────────────────────────────────────────────────────
#  AUTENTICACIÓN HÍBRIDA ROBUSTA (TOKEN / BEARER)
# ─────────────────────────────────────────────────────────────────
class TokenAuthentication(BaseAuthentication):
    def authenticate(self, request):
        auth_header = request.headers.get('Authorization')
        if not auth_header: return None
        parts = auth_header.split()
        if len(parts) == 2: token_key = parts[1]
        elif len(parts) == 1: token_key = parts[0]
        else: return None
        try:
            token = Token.objects.select_related('usuario').get(key=token_key)
            if not token.usuario.activo: raise AuthenticationFailed('Usuario inactivo.')
            return (token.usuario, token)
        except Token.DoesNotExist:
            raise AuthenticationFailed('Token inválido.')

# ─────────────────────────────────────────────
#  AUTH
# ─────────────────────────────────────────────

@api_view(['GET', 'POST']) 
@permission_classes([AllowAny])
def login_view(request):
    if request.method == 'POST':
        username = request.data.get('username') or request.POST.get('username')
        password = request.data.get('password') or request.POST.get('password')
        
        user = authenticate(username=username, password=password)
        if user is not None:
            auth_login(request, user)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.META.get('HTTP_ACCEPT', ''):
                return JsonResponse({'success': True, 'redirect': '/api/panel/dashboard/'})
            return redirect('panel_dashboard')
        
        if 'application/json' in request.META.get('HTTP_ACCEPT', ''):
            return JsonResponse({'error': 'Credenciales inválidas'}, status=400)
        return render(request, 'tickets/auth/login.html', {'error': 'Credenciales inválidas'})

    return render(request, 'tickets/auth/login.html')

@api_view(['GET', 'POST']) 
@permission_classes([AllowAny]) 
def logout_view(request):
    if request.user.is_authenticated and hasattr(request.user, 'auth_token'):
        Token.objects.filter(user=request.user).delete()
    
    django_logout(request)
    
    if request.method == 'GET' or 'text/html' in request.META.get('HTTP_ACCEPT', ''):
        return redirect('login') 
        
    return Response({"success": "Sesión cerrada correctamente"}, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def me_view(request):
    return Response(UsuarioSerializer(request.user).data)

# ─────────────────────────────────────────────
#  SLA LOGS
# ─────────────────────────────────────────────

def _handle_state_change(ticket, old_estado, new_estado, user):
    now = timezone.now()
    update_fields = []
    
    if ticket.usuario_asignado:
        update_fields.append('usuario_asignado')
        if not ticket.fecha_asignacion:
            ticket.fecha_asignacion = now
            update_fields.append('fecha_asignacion')
    else:
        update_fields.append('usuario_asignado')

    if old_estado and not ticket.fecha_primera_respuesta:
        ticket.fecha_primera_respuesta = now
        update_fields.append('fecha_primera_respuesta')

    if old_estado and old_estado.pausa_sla:
        open_log = TicketTimeLog.objects.filter(ticket=ticket, fecha_fin__isnull=True).first()
        if open_log:
            open_log.fecha_fin = now
            open_log.save()
            ticket.tiempo_pausa_minutos = sum(log.duracion_minutos for log in TicketTimeLog.objects.filter(ticket=ticket, duracion_minutos__isnull=False))
            update_fields.append('tiempo_pausa_minutos')
            
    if new_estado and new_estado.pausa_sla:
        TicketTimeLog.objects.create(ticket=ticket, estado_pausa=new_estado.nombre, fecha_inicio=now)
        
    if new_estado and 'resuelto' in new_estado.nombre.lower():
        ticket.fecha_resolucion = now
        update_fields.append('fecha_resolucion')
        
        if ticket.fecha_creacion:
            tiempo_bruto_minutos = int((now - ticket.fecha_creacion).total_seconds() / 60)
            pausas = ticket.tiempo_pausa_minutos or 0
            ticket.tiempo_atencion_minutos = max(0, tiempo_bruto_minutos - pausas)
            update_fields.append('tiempo_atencion_minutos')

    elif new_estado and 'cerrado' in new_estado.nombre.lower():
        ticket.fecha_cierre = now
        update_fields.append('fecha_cierre')
        
    elif new_estado and new_estado.es_estado_cierre:
        if not ticket.fecha_resolucion:
            ticket.fecha_resolucion = now
            update_fields.append('fecha_resolucion')
        if not ticket.fecha_cierre:
            ticket.fecha_cierre = now
            update_fields.append('fecha_cierre')
            
    if 'estado' not in update_fields:
        update_fields.append('estado')

    ticket.save(update_fields=update_fields) if update_fields else ticket.save()
        
    contenido_sistema = f"Estado cambiado a '{new_estado.nombre if new_estado else '—'}'"
    
    ChatterEntry.objects.create(
        ticket=ticket, 
        tipo='sistema', 
        autor=user, 
        estado_anterior=old_estado.nombre if old_estado else None, 
        estado_nuevo=new_estado.nombre if new_estado else None, 
        contenido=contenido_sistema
    )

    lista_correos = []
    if ticket.usuario_reporta and getattr(ticket.usuario_reporta, 'correo_electronico', None):
        lista_correos.append(ticket.usuario_reporta.correo_electronico)
    if ticket.usuario_asignado and getattr(ticket.usuario_asignado, 'correo_electronico', None):
        lista_correos.append(ticket.usuario_asignado.correo_electronico)
    if getattr(ticket, 'correos_seguimiento', None):
        adicionales = [c.strip() for c in ticket.correos_seguimiento.split(',') if c.strip()]
        lista_correos.extend(adicionales)

    lista_correos = list(set(lista_correos))

    if lista_correos:
        folio_ticket = getattr(ticket, 'folio', ticket.id)
        titulo_ticket = getattr(ticket, 'titulo', 'Soporte Técnico')
        nombre_operador = user.nombre_completo if (user and hasattr(user, 'nombre_completo') and user.nombre_completo) else "Sistema"
        
        asunto = f"⚙️ Cambio de Estado en Ticket #{folio_ticket} - {titulo_ticket}"
        html_contenido = f"""
        <div style="font-family: sans-serif; max-width: 600px; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;">
            <div style="background-color: #1e293b; padding: 20px; color: #38bdf8; font-weight: bold; font-size: 16px;">
                ⚙️ Actualización Automática de Sistema - Ticket #{folio_ticket}
            </div>
            <div style="padding: 20px; font-size: 13px; line-height: 1.6; color: #334155;">
                <p>Se ha registrado un movimiento automático de control por el operador <strong>{nombre_operador}</strong>:</p>
                <div style="margin: 15px 0; padding: 12px; border-left: 4px solid #38bdf8; background-color: #f8fafc; font-weight: 500;">
                    {contenido_sistema}
                </div>
                <p>Puedes verificar los tiempos de atención, las matrices de criticidad y el SLA ingresando al panel de control.</p>
            </div>
        </div>
        """
        hilo_sistema = threading.Thread(
            target=_tarea_enviar_correo_async,
            args=(asunto, html_contenido, settings.DEFAULT_FROM_EMAIL, lista_correos)
        )
        hilo_sistema.daemon = True
        hilo_sistema.start()

# ─────────────────────────────────────────────
#  FUNCIÓN REUTILIZABLE DE LIMPIEZA DE FECHAS
# ─────────────────────────────────────────────
def _clean_view_date_string(date_str):
    if not date_str:
        return "2026-06-25T00:00:00Z"
    if '-' in date_str and date_str.count('-') == 3:
        date_str = date_str.rsplit('-', 1)[0]
    elif '+' in date_str:
        date_str = date_str.rsplit('+', 1)[0]
    if date_str.endswith('Z'):
        date_str = date_str[:-1]
    return date_str + "Z"

# ─────────────────────────────────────────────
#  VIEWSETS
# ─────────────────────────────────────────────

class TicketViewSet(viewsets.ModelViewSet):
    queryset = Ticket.objects.select_related('sistema', 'modulo', 'prioridad', 'estado', 'categoria', 'usuario_reporta', 'usuario_asignado').all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['estado', 'prioridad', 'categoria', 'sistema', 'modulo', 'usuario_asignado', 'usuario_reporta']
    search_fields = ['folio', 'titulo', 'descripcion', 'codigo_error']
    ordering_fields = ['fecha_creacion', 'prioridad__orden', 'estado__orden']
    ordering = ['-fecha_creacion']
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def retrieve(self, request, pk=None, *args, **kwargs):
        try:
            instance = Ticket.objects.select_related(
                'sistema', 'modulo', 'prioridad', 'estado', 'categoria', 'usuario_reporta', 'usuario_asignado'
            ).get(pk=pk)
        except Ticket.DoesNotExist:
            return Response({'detail': f'Ticket {pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(instance)
        data = serializer.data

        data['sistema'] = instance.sistema.id if instance.sistema else None
        data['modulo'] = instance.modulo.id if instance.modulo else None
        data['prioridad'] = instance.prioridad.id if instance.prioridad else None
        data['estado'] = instance.estado.id if instance.estado else None
        data['categoria'] = instance.categoria.id if instance.categoria else None
        data['usuario_reporta'] = instance.usuario_reporta.id if instance.usuario_reporta else None
        data['usuario_asignado'] = instance.usuario_asignado.id if instance.usuario_asignado else None

        data['sistema_id'] = data['sistema']
        data['modulo_id'] = data['modulo']
        data['prioridad_id'] = data['prioridad']
        data['estado_id'] = data['estado']
        data['categoria_id'] = data['categoria']
        data['usuario_reporta_id'] = data['usuario_reporta']
        data['usuario_asignado_id'] = data['usuario_asignado']

        data['sistema_nombre'] = instance.sistema.nombre if instance.sistema else "—"
        data['modulo_nombre'] = instance.modulo.nombre if instance.modulo else "—"
        data['prioridad_nombre'] = instance.prioridad.nombre if instance.prioridad else "—"
        data['prioridad_color'] = instance.prioridad.color if instance.prioridad else ""
        data['estado_nombre'] = instance.estado.nombre if instance.estado else "—"
        data['estado_color'] = instance.estado.color if instance.estado else ""
        data['categoria_nombre'] = instance.categoria.nombre if instance.categoria else "—"
        data['usuario_reporta_nombre'] = instance.usuario_reporta.nombre_completo if instance.usuario_reporta else "—"
        data['usuario_asignado_nombre'] = instance.usuario_asignado.nombre_completo if instance.usuario_asignado else "Sin asignar"

        def _clean_date(dt):
            if not dt:
                return "2026-06-25T00:00:00Z"
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

        data['fecha_creacion'] = _clean_date(instance.fecha_creacion)
        data['fecha_asignacion'] = _clean_date(instance.fecha_asignacion) if instance.fecha_asignacion else data['fecha_creacion']
        data['fecha_primera_respuesta'] = _clean_date(instance.fecha_primera_respuesta) if instance.fecha_primera_respuesta else data['fecha_creacion']
        data['fecha_resolucion'] = _clean_date(instance.fecha_resolucion) if instance.fecha_resolucion else data['fecha_creacion']
        data['fecha_cierre'] = _clean_date(instance.fecha_cierre) if instance.fecha_cierre else data['fecha_creacion']

        return Response(data)

    def create(self, request, *args, **kwargs):
        data = request.data.get('data') if 'data' in request.data else request.data
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        ticket = serializer.save()
        if ticket.estado and ticket.estado.pausa_sla: TicketTimeLog.objects.create(ticket=ticket, estado_pausa=ticket.estado.nombre, fecha_inicio=timezone.now())
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_queryset(self):
        qs = super().get_queryset()
        incluir_archivados = self.request.query_params.get('ver_archivados', 'false') == 'true'
        if not incluir_archivados:
            qs = qs.filter(archivado=False)
            
        vista = self.request.query_params.get('vista')
        if not vista or vista == 'todos': return qs
        now = timezone.now()
        if vista == 'abiertos': return qs.filter(estado__es_estado_cierre=False)
        if vista == 'en_proceso': return qs.filter(estado__es_estado_cierre=False, usuario_asignado__isnull=False)
        if vista == 'resueltos': return qs.filter(estado__es_estado_cierre=True, fecha_cierre__isnull=True)
        if vista == 'cerrados': return qs.filter(estado__es_estado_cierre=True)
        if vista == 'hoy': return qs.filter(fecha_creacion__date=now.date())
        return qs

    def get_serializer_class(self):
        if self.action == 'create': return TicketInputSerializer
        if self.action in ['partial_update', 'update']: return TicketUpdateSerializer
        return TicketSerializer


class SistemaViewSet(viewsets.ModelViewSet):
    queryset = Sistema.objects.all()
    serializer_class = SistemaSerializer
    pagination_class = None

class ModuloViewSet(viewsets.ModelViewSet):
    queryset = Modulo.objects.all()
    serializer_class = ModuloSerializer
    pagination_class = None

class DocumentoViewSet(viewsets.ModelViewSet):
    queryset = Documento.objects.all()
    serializer_class = DocumentoSerializer
    pagination_class = None

class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    pagination_class = None
    def get_serializer_class(self):
        if self.action in ['partial_update', 'update']: return UsuarioUpdateSerializer
        return UsuarioSerializer

class PrioridadViewSet(viewsets.ModelViewSet):
    queryset = Prioridad.objects.all()
    serializer_class = PrioridadSerializer
    pagination_class = None

class EstadoViewSet(viewsets.ModelViewSet):
    queryset = Estado.objects.all()
    serializer_class = EstadoSerializer
    pagination_class = None

class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    pagination_class = None

class ConocimientoViewSet(viewsets.ModelViewSet):
    queryset = ConocimientoEntry.objects.all()
    serializer_class = ConocimientoSerializer
    pagination_class = None

# ─────────────────────────────────────────────
#  REPORTES CORE
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([AllowAny])
def reporte_resumen(request):
    now = timezone.now()
    total = Ticket.objects.count()
    abiertos = Ticket.objects.filter(estado__es_estado_cierre=False).count()
    en_proceso = Ticket.objects.filter(estado__es_estado_cierre=False, usuario_asignado__isnull=False).count()
    resueltos = Ticket.objects.filter(estado__es_estado_cierre=True, fecha_cierre__isnull=True).count()
    cerrados = Ticket.objects.filter(estado__es_estado_cierre=True).count()
    avg_res = Ticket.objects.filter(tiempo_atencion_minutos__isnull=False).aggregate(avg=Avg('tiempo_atencion_minutos'))['avg'] or 0
    return Response({
        'total_tickets': total, 'abiertos': abiertos, 'en_proceso': en_proceso, 'resueltos': resueltos, 'cerrados': cerrados,
        'vencidos': 0, 'tickets_hoy': Ticket.objects.filter(fecha_creacion__date=now.date()).count(), 'tickets_semana': 0,
        'promedio_resolucion_horas': round(avg_res / 60, 2), 'satisfaccion_promedio': 5.0, 'porcentaje_sla_cumplido': 100.0
    })

@api_view(['GET'])
def reporte_por_sistema(request):
    try:
        data = Ticket.objects.values('sistema__id', 'sistema__nombre').annotate(total=Count('id'))
        result = []
        for r in data:
            result.append({
                'id': r['sistema__id'] or 0,
                'nombre': r['sistema__nombre'] or 'Sin sistema',
                'total': r['total'] or 0
            })
        return Response(list(result), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def reporte_por_estado(request):
    try:
        data = Ticket.objects.values('estado__id', 'estado__nombre', 'estado__color').annotate(total=Count('id'))
        result = []
        for r in data:
            result.append({
                'id': r['estado__id'] or 0,
                'nombre': r['estado__nombre'] or 'Sin estado',
                'total': r['total'] or 0,
                'color': r['estado__color'] or '#gray'
            })
        return Response(list(result), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def reporte_por_prioridad(request):
    try:
        data = Ticket.objects.values('prioridad__id', 'prioridad__nombre', 'prioridad__color').annotate(total=Count('id'))
        result = []
        for r in data:
            result.append({
                'id': r['prioridad__id'] or 0,
                'nombre': r['prioridad__nombre'] or 'Sin prioridad',
                'total': r['total'] or 0,
                'color': r['prioridad__color'] or '#gray'
            })
        return Response(list(result), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def reporte_sla(request):
    try:
        return Response({
            'promedio_primera_respuesta_horas': 0.0,
            'promedio_resolucion_horas': 0.0,
            'cumplimiento_sla_porcentaje': 100.0,
            'por_prioridad': []  
        }, status=status.HTTP_200_OK)
    except Exception:
        return Response({'por_prioridad': []}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def reporte_tendencias(request):
    try:
        result = []
        for i in range(29, -1, -1):
            day = (timezone.now() - timedelta(days=i)).date()
            result.append({
                'fecha': str(day), 
                'total': Ticket.objects.filter(fecha_creacion__date=day).count(), 
                'resueltos': Ticket.objects.filter(fecha_cierre__date=day).count()
            })
        return Response(list(result), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def reporte_por_region(request):
    try:
        return Response([], status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def reporte_tickets(request): 
    try:
        qs = Ticket.objects.select_related(
            'sistema', 'modulo', 'prioridad', 'estado', 'categoria', 'usuario_reporta', 'usuario_asignado'
        ).all().order_by('-fecha_creacion')[:100]
        
        result = []
        for instance in qs:
            row = {
                'id': instance.id,
                'folio': instance.folio or "—",
                'titulo': instance.titulo or "—",
                'descripcion': instance.descripcion or "",
                'codigo_error': instance.codigo_error or "",
                'tiempo_atencion_minutos': instance.tiempo_atencion_minutos or 0,
                'tiempo_pausa_minutos': instance.tiempo_pausa_minutos or 0,
                'sistema_id': instance.sistema.id if instance.sistema else None,
                'modulo_id': instance.modulo.id if instance.modulo else None,
                'prioridad_id': instance.prioridad.id if instance.prioridad else None,
                'estado_id': instance.estado.id if instance.estado else None,
                'categoria_id': instance.categoria.id if instance.categoria else None,
                'sistema_nombre': instance.sistema.nombre if instance.sistema else "—",
                'modulo_nombre': instance.modulo.nombre if instance.modulo else "—",
                'prioridad_nombre': instance.prioridad.nombre if instance.prioridad else "—",
                'prioridad_color': instance.prioridad.color if instance.prioridad else "",
                'estado_nombre': instance.estado.nombre if instance.estado else "—",
                'estado_color': instance.estado.color if instance.estado else "",
                'categoria_nombre': instance.categoria.nombre if instance.categoria else "—",
                'usuario_reporta_nombre': instance.usuario_reporta.nombre_completo if instance.usuario_reporta else "—",
                'usuario_asignado_nombre': instance.usuario_asignado.nombre_completo if instance.usuario_asignado else "Sin asignar",
            }
            
            def _clean_date(dt):
                if not dt:
                    return "2026-06-25T00:00:00Z"
                return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
                
            row['fecha_creacion'] = _clean_date(instance.fecha_creacion)
            row['fecha_asignacion'] = _clean_date(instance.fecha_asignacion) if instance.fecha_asignacion else row['fecha_creacion']
            row['fecha_primera_respuesta'] = _clean_date(instance.fecha_primera_respuesta) if instance.fecha_primera_respuesta else row['fecha_creacion']
            row['fecha_resolucion'] = _clean_date(instance.fecha_resolucion) if instance.fecha_resolucion else row['fecha_creacion']
            row['fecha_cierre'] = _clean_date(instance.fecha_cierre) if instance.fecha_cierre else row['fecha_creacion']
            
            result.append(row)
            
        return Response(list(result), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def actividad_reciente(request): return Response([])

# ─────────────────────────────────────────────────────────────────
#  ENDPOINTS DE COMPATIBILIDAD (CRUD)
# ─────────────────────────────────────────────────────────────────

@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_create_usuario(request):
    payload = request.data.get('data') if 'data' in request.data else request.data
    serializer = UsuarioInputSerializer(data=payload if payload else request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['POST', 'PUT', 'PATCH', 'GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_update_usuario(request, pk=None):
    payload = request.data.get('data') if 'data' in request.data else request.data
    if payload is None: payload = request.data
    custom_data = payload.copy() if hasattr(payload, 'copy') else dict(payload)
    if 'estado' in custom_data:
        custom_data['activo'] = custom_data['estado'] in ['Activo', 'activo', True, 'true', 'True', 1, '1']
    usuario_id = pk or custom_data.get('id') or request.query_params.get('id')
    try:
        usuario = Usuario.objects.get(id=usuario_id)
    except Usuario.DoesNotExist:
        return Response({'detail': 'No encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    serializer = UsuarioUpdateSerializer(usuario, data=custom_data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_delete_usuario(request, pk=None):
    payload = request.data.get('data') if 'data' in request.data else request.data
    if payload is None: payload = request.data
    usuario_id = pk or payload.get('id')
    Usuario.objects.filter(id=usuario_id).delete()
    return Response({'detail': 'Eliminado.'}, status=status.HTTP_200_OK)

@api_view(['POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_delete_modulo(request, pk=None):
    payload = request.data.get('data') if 'data' in request.data else request.data
    if payload is None: payload = request.data
    modulo_id = pk or payload.get('id')
    Modulo.objects.filter(id=modulo_id).delete()
    return Response({'detail': 'Módulo eliminado.'}, status=status.HTTP_200_OK)

@api_view(['POST', 'DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_delete_conocimiento(request, pk=None):
    payload = request.data.get('data') if 'data' in request.data else request.data
    if payload is None: payload = request.data
    entry_id = pk or payload.get('id')
    ConocimientoEntry.objects.filter(id=entry_id).delete()
    return Response({'detail': 'Entrada eliminada.'}, status=status.HTTP_200_OK)

@api_view(['POST', 'GET'])
def compat_create_ticket(request):
    serializer = TicketInputSerializer(data=request.data.get('data', request.data))
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['POST', 'GET'])
def compat_create_modulo(request):
    serializer = ModuloSerializer(data=request.data.get('data', request.data))
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['POST', 'GET'])
def compat_create_conocimiento(request):
    serializer = ConocimientoSerializer(data=request.data.get('data', request.data))
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
def compat_chatter_list(request):
    tid = request.query_params.get('ticket') or request.query_params.get('ticket_id') or request.query_params.get('id')
    if not tid: return Response([], status=status.HTTP_200_OK)
    try:
        queryset = ChatterEntry.objects.filter(ticket_id=int(tid)).order_by('fecha_creacion')
        serializer = ChatterEntrySerializer(queryset, many=True)
        return Response(list(serializer.data), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['GET'])
def compat_timelogs_list(request):
    tid = request.query_params.get('ticket') or request.query_params.get('ticket_id') or request.query_params.get('id')
    if not tid: return Response([], status=status.HTTP_200_OK)
    try:
        queryset = TicketTimeLog.objects.filter(ticket_id=int(tid)).order_by('fecha_inicio')
        serializer = TimeLogSerializer(queryset, many=True)
        return Response(list(serializer.data), status=status.HTTP_200_OK)
    except Exception:
        return Response([], status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_add_chatter(request):
    payload = request.data.get('data', request.data)
    ticket_id = payload.get('ticket_id') or payload.get('ticket')
    contenido = payload.get('contenido')
    if not ticket_id or not contenido: return Response({'detail': 'Faltan parámetros.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        ticket = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return Response({'detail': 'Ticket no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    entry = ChatterEntry.objects.create(ticket=ticket, tipo='comentario', autor=request.user, contenido=contenido)
    return Response(ChatterEntrySerializer(entry).data, status=status.HTTP_201_CREATED)

@api_view(['POST', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_update_ticket(request):
    payload = request.data.get('data', request.data)
    ticket_id = payload.get('id') or request.query_params.get('id')
    try:
        ticket = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return Response({'detail': 'Ticket no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    old_est = ticket.estado
    serializer = TicketUpdateSerializer(ticket, data=payload, partial=True)
    serializer.is_valid(raise_exception=True)
    ticket_upd = serializer.save()
    if old_est != ticket_upd.estado:
        _handle_state_change(ticket_upd, old_est, ticket_upd.estado, request.user)
    return Response(TicketSerializer(ticket_upd).data)

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def compat_ticket_detail(request, pk):
    try:
        ticket = Ticket.objects.select_related('sistema', 'modulo', 'prioridad', 'estado', 'categoria', 'usuario_reporta', 'usuario_asignado').get(pk=pk)
    except Ticket.DoesNotExist:
        return Response({'detail': 'No encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = TicketSerializer(ticket)
        data = serializer.data
        base_date = _clean_view_date_string(data.get('fecha_creacion'))
        data['fecha_creacion'] = _clean_view_date_string(data.get('fecha_creacion'))
        data['fecha_asignacion'] = _clean_view_date_string(data.get('fecha_asignacion')) if data.get('fecha_asignacion') else base_date
        data['fecha_primera_respuesta'] = _clean_view_date_string(data.get('fecha_primera_respuesta')) if data.get('fecha_primera_respuesta') else base_date
        data['fecha_resolucion'] = _clean_view_date_string(data.get('fecha_resolucion')) if data.get('fecha_resolucion') else base_date
        data['fecha_cierre'] = _clean_view_date_string(data.get('fecha_cierre')) if data.get('fecha_cierre') else base_date
        return Response(data)

    elif request.method in ['PUT', 'PATCH']:
        payload = request.data.get('data', request.data)
        old_est = ticket.estado
        serializer = TicketUpdateSerializer(ticket, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        ticket_upd = serializer.save()
        if old_est != ticket_upd.estado: 
            _handle_state_change(ticket_upd, old_est, ticket_upd.estado, request.user)
        return_serializer = TicketSerializer(ticket_upd)
        return_data = return_serializer.data
        base_date = _clean_view_date_string(return_data.get('fecha_creacion'))
        return_data['fecha_creacion'] = _clean_view_date_string(return_data.get('fecha_creacion'))
        return_data['fecha_asignacion'] = _clean_view_date_string(return_data.get('fecha_asignacion')) if return_data.get('fecha_asignacion') else base_date
        return_data['fecha_cierre'] = _clean_view_date_string(return_data.get('fecha_cierre')) if return_data.get('fecha_cierre') else base_date
        return Response(return_data)

    elif request.method == 'DELETE':
        ticket.delete()
        return Response({'detail': 'Eliminado.'}, status=status.HTTP_200_OK)

# ─────────────────────────────────────────────
#  VISTAS HTMX E INTERNAS
# ─────────────────────────────────────────────

@login_required
def panel_tickets_list(request):
    filtrar = request.GET.get('filtrar', '')
    query = request.GET.get('q', '').strip()
    ordering = request.GET.get('ordering', '-fecha_creacion')
    asignado_id = request.GET.get('asignado_id')
    prioridad_id = request.GET.get('prioridad_id')
    estado_id = request.GET.get('estado_id')
    impacto = request.GET.get('impacto')
    archivado_filtro = request.GET.get('archivado', 'false')
    
    qs = Ticket.objects.select_related('sistema', 'modulo', 'prioridad', 'estado', 'usuario_asignado')

    if archivado_filtro == 'true': qs = qs.filter(archivado=True)
    elif archivado_filtro == 'false': qs = qs.filter(archivado=False)

    titulo_panel = "Panel Global de Tickets"
    if filtrar == 'pendientes':
        qs = qs.exclude(estado__nombre__icontains='cerrado').exclude(estado__nombre__icontains='resuelto')
        titulo_panel = "Tickets Pendientes (Abiertos)"
    elif filtrar == 'resueltos':
        qs = qs.filter(Q(estado__nombre__icontains='cerrado') | Q(estado__nombre__icontains='resuelto'))
        titulo_panel = "Tickets Resueltos / Cerrados"

    if query:
        qs = qs.filter(Q(folio__icontains=query) | Q(titulo__icontains=query) | Q(descripcion__icontains=query) | Q(usuario_asignado__nombre_completo__icontains=query))

    if asignado_id: qs = qs.filter(usuario_asignado_id=asignado_id)
    if prioridad_id: qs = qs.filter(prioridad_id=prioridad_id)
    if estado_id: qs = qs.filter(estado_id=estado_id)
    if impacto: qs = qs.filter(impacto_proceso=impacto)

    campos_permitidos = [
        'folio', '-folio', 'titulo', '-titulo', 'usuario_asignado__nombre_completo', '-usuario_asignado__nombre_completo',
        'prioridad__orden', '-prioridad__orden', 'estado__orden', '-estado__orden', 'impacto_proceso', '-impacto_proceso', 'fecha_creacion', '-fecha_creacion'
    ]
    qs = qs.order_by(ordering) if ordering in campos_permitidos else qs.order_by('-fecha_creacion')
    tickets = qs[:100]
    
    context = {
        'tickets': tickets, 'titulo_panel': titulo_panel, 'current_ordering': ordering,
        'next_folio': '-folio' if ordering == 'folio' else 'folio',
        'next_titulo': '-titulo' if ordering == 'titulo' else 'titulo',
        'next_asignado': '-usuario_asignado__nombre_completo' if ordering == 'usuario_asignado__nombre_completo' else 'usuario_asignado__nombre_completo',
        'next_prioridad': '-prioridad__orden' if ordering == 'prioridad__orden' else 'prioridad__orden',
        'next_estado': '-estado__orden' if ordering == 'estado__orden' else 'estado__orden',
        'next_impacto': '-impacto_proceso' if ordering == 'impacto_proceso' else 'impacto_proceso',
        'estados': Estado.objects.all().order_by('orden'), 'prioridades': Prioridad.objects.all(),
        'tecnicos': Usuario.objects.filter(rol='tecnico') or Usuario.objects.filter(is_staff=True) or Usuario.objects.all(),
        'current_archivado': archivado_filtro,
    }
    if request.headers.get('HX-Request'): return render(request, 'tickets/partials/tickets_render_search.html', context)
    return render(request, 'tickets/list.html', context)

@login_required
def panel_tickets_exportar_excel(request):
    query = request.GET.get('q', '').strip()
    ordering = request.GET.get('ordering', '-fecha_creacion')
    asignado_id = request.GET.get('asignado_id')
    prioridad_id = request.GET.get('prioridad_id')
    estado_id = request.GET.get('estado_id')
    impacto = request.GET.get('impacto')
    archivado_filtro = request.GET.get('archivado', 'false')

    qs = Ticket.objects.select_related('sistema', 'modulo', 'estado', 'prioridad', 'usuario_asignado').all()
    if query: qs = qs.filter(Q(folio__icontains=query) | Q(titulo__icontains=query) | Q(descripcion__icontains=query) | Q(usuario_asignado__nombre_completo__icontains=query))
    if asignado_id: qs = qs.filter(usuario_asignado_id=asignado_id)
    if prioridad_id: qs = qs.filter(prioridad_id=prioridad_id)
    if estado_id: qs = qs.filter(estado_id=estado_id)
    if impacto: qs = qs.filter(impacto_proceso=impacto)
    if archivado_filtro == 'true': qs = qs.filter(archivado=True)
    elif archivado_filtro == 'false': qs = qs.filter(archivado=False)

    campos_permitidos = [
        'folio', '-folio', 'titulo', '-titulo', 'usuario_asignado__nombre_completo', '-usuario_asignado__nombre_completo',
        'prioridad__orden', '-prioridad__orden', 'estado__orden', '-estado__orden', 'impacto_proceso', '-impacto_proceso', 'fecha_creacion', '-fecha_creacion'
    ]
    qs = qs.order_by(ordering) if ordering in campos_permitidos else qs.order_by('-fecha_creacion')

    response = HttpResponse(content_type='text/csv; charset=windows-1252')
    response['Content-Disposition'] = 'attachment; filename="reporte_tickets_filtrado.csv"'
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Folio', 'Titulo', 'Sistema', 'Modulo', 'Prioridad', 'Estado', 'Impacto', 'Asignado A', 'Fecha Creacion'])

    for tk in qs:
        impacto_txt = tk.get_impacto_proceso_display() if hasattr(tk, 'get_impacto_proceso_display') else (tk.impacto_proceso or 'Funcional')
        writer.writerow([
            tk.folio, str(tk.titulo).encode('windows-1252', 'replace').decode('windows-1252'),
            str(tk.sistema.nombre if tk.sistema else '—').encode('windows-1252', 'replace').decode('windows-1252'),
            str(tk.modulo.nombre if tk.modulo else '—').encode('windows-1252', 'replace').decode('windows-1252'),
            tk.prioridad.nombre if tk.prioridad else '—', tk.estado.nombre if tk.estado else '—', impacto_txt,
            str(tk.usuario_asignado.nombre_completo if tk.usuario_asignado else 'Sin Asignar').encode('windows-1252', 'replace').decode('windows-1252'),
            tk.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])
    return response

@login_required
def panel_ticket_chatter(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.method == "POST":
        contenido = request.POST.get("contenido", "").strip()
        if contenido:
            nueva_nota = ChatterEntry.objects.create(ticket=ticket, tipo='comentario', contenido=contenido, autor=request.user)
            lista_correos = []
            if ticket.usuario_reporta and hasattr(ticket.usuario_reporta, 'correo_electronico') and ticket.usuario_reporta.correo_electronico:
                lista_correos.append(ticket.usuario_reporta.correo_electronico)
            if ticket.usuario_asignado and hasattr(ticket.usuario_asignado, 'correo_electronico') and ticket.usuario_asignado.correo_electronico:
                lista_correos.append(ticket.usuario_asignado.correo_electronico)
            if ticket.correos_seguimiento:
                adicionales = [c.strip() for c in ticket.correos_seguimiento.split(',') if c.strip()]
                lista_correos.extend(adicionales)

            lista_correos = list(set(lista_correos))
            if lista_correos:
                folio_ticket = getattr(ticket, 'folio', ticket.id)
                titulo_ticket = getattr(ticket, 'titulo', 'Soporte Técnico')
                nombre_remitente = getattr(request.user, 'nombre_completo', None) or getattr(request.user, 'correo_electronico', 'Soporte')
                asunto = f"🔔 Actualización en Ticket #{folio_ticket} - {titulo_ticket}"
                html_contenido = f"""
                <div style="font-family: sans-serif; max-width: 600px; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;">
                    <div style="background-color: #0f172a; padding: 20px; color: #fbbf24; font-weight: bold; font-size: 16px;">
                        🔔 Nueva Nota en Ticket #{folio_ticket}
                    </div>
                    <div style="padding: 20px; font-size: 13px; line-height: 1.6; color: #334155;">
                        <p>El usuario <strong>{nombre_remitente}</strong> ha agregado una nueva actualización al Historial de Notas:</p>
                        <blockquote style="margin: 15px 0; padding: 10px 15px; border-left: 4px solid #fbbf24; background-color: #f8fafc; font-style: italic;">
                            "{contenido}"
                        </blockquote>
                        <p>Puedes revisar el estatus completo ingresando a la plataforma institucional.</p>
                    </div>
                </div>
                """
                hilo_correo = threading.Thread(target=_tarea_enviar_correo_async, args=(asunto, html_contenido, settings.DEFAULT_FROM_EMAIL, lista_correos))
                hilo_correo.daemon = True
                hilo_correo.start()

    notas = ticket.chatter.all().order_by('-fecha_creacion')
    html_output = ""
    for nota in notas:
        is_sistema = (nota.tipo == 'sistema')
        bg_class = "bg-slate-50/60 dark:bg-[#13161c]/40 border-slate-100 dark:border-slate-800 text-slate-500 dark:text-orange-500 font-medium" if is_sistema else "bg-white dark:bg-[#252a34] border-slate-200 dark:border-slate-800 text-slate-700 dark:text-slate-200"
        autor_name = "🤖 Sistema" if is_sistema else (nota.autor.nombre_completo if nota.autor else "Usuario")
        fecha = nota.fecha_creacion.strftime("%d/%m/%Y %H:%M")
        html_output += f"""
        <div class="p-3 rounded-lg border text-xs transition duration-150 {bg_class}">
            <div class="flex items-center justify-between font-semibold mb-1 text-[11px]">
                <span>{autor_name}</span>
                <span class="text-slate-400 dark:text-slate-500 font-mono text-[10px]">{fecha}</span>
            </div>
            <p class="whitespace-pre-wrap leading-relaxed pl-0.5">{nota.contenido}</p>
        </div>
        """
    if not html_output: html_output = '<div class="text-center py-4 text-xs text-slate-400 dark:text-orange-500/60 italic">No hay notas registradas todavía.</div>'
    return HttpResponse(html_output)

@login_required
def panel_dashboard(request):
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    hoy = timezone.now().date()
    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else date(hoy.year, 1, 1)
    fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else hoy

    tickets_filtrados = Ticket.objects.filter(fecha_creacion__date__range=[fecha_inicio, fecha_fin], archivado=False)
    total_tickets = tickets_filtrados.count()
    pendientes = tickets_filtrados.filter(~Q(estado__es_estado_cierre=True)).count()
    resueltos = tickets_filtrados.filter(estado__es_estado_cierre=True).count()

    tickets_resueltos = tickets_filtrados.filter(estado__es_estado_cierre=True)
    resueltos_count = tickets_resueltos.count()
    tickets_cumplieron_sla = 0

    sla_agentes_dict = {}
    sla_sistemas_dict = {}

    for tk in tickets_resueltos:
        prioridad_name = str(tk.prioridad.nombre).strip().lower() if tk.prioridad else 'bajo'
        impacto_name = str(tk.impacto_proceso).strip().lower() if tk.impacto_proceso else 'baja'
        agente_name = tk.usuario_asignado.nombre_completo if tk.usuario_asignado else "Sin Asignar"
        sistema_name = tk.sistema.nombre if tk.sistema else "General"
        
        if agente_name not in sla_agentes_dict: sla_agentes_dict[agente_name] = {'total_cerrados': 0, 'cumplidos': 0}
        if sistema_name not in sla_sistemas_dict: sla_sistemas_dict[sistema_name] = {'total_cerrados': 0, 'cumplidos': 0, 'total_minutos': 0}

        limite_horas = 48
        if 'alto' in prioridad_name or 'critica' in prioridad_name or 'crítica' in prioridad_name:
            if 'alta' in impacto_name or 'caida' in impacto_name or 'caída' in impacto_name: limite_horas = 2
            elif 'media' in impacto_name or 'parcial' in impacto_name: limite_horas = 4
            else: limite_horas = 8
        elif 'medio' in prioridad_name:
            if 'alta' in impacto_name or 'caida' in impacto_name or 'caída' in impacto_name: limite_horas = 4
            elif 'media' in impacto_name or 'parcial' in impacto_name: limite_horas = 12
            else: limite_horas = 24
        else:
            if 'alta' in impacto_name or 'caida' in impacto_name or 'caída' in impacto_name: limite_horas = 8
            elif 'media' in impacto_name or 'parcial' in impacto_name: limite_horas = 24
            else: limite_horas = 48

        limite_minutos_dinamico = limite_horas * 60
        tiempo_atencion = tk.tiempo_atencion_minutos or 0
        
        sla_agentes_dict[agente_name]['total_cerrados'] += 1
        sla_sistemas_dict[sistema_name]['total_cerrados'] += 1
        sla_sistemas_dict[sistema_name]['total_minutos'] += tiempo_atencion
        
        if tiempo_atencion <= limite_minutos_dinamico:
            tickets_cumplieron_sla += 1
            sla_agentes_dict[agente_name]['cumplidos'] += 1
            sla_sistemas_dict[sistema_name]['cumplidos'] += 1

    sla_porcentaje = int((tickets_cumplieron_sla / resueltos_count) * 100) if resueltos_count > 0 else 100

    tendencias_data = tickets_filtrados.annotate(dia=TruncDate('fecha_creacion')).values('dia').annotate(total=Count('id')).order_by('dia')
    tendencias_labels = [item['dia'].strftime('%Y-%m-%d') for item in tendencias_data if item['dia']]
    tendencias_valores = [item['total'] for item in tendencias_data]

    estados_data = tickets_filtrados.values('estado__nombre').annotate(total=Count('id')).order_by('-total')
    estados_labels = [item['estado__nombre'] or "Sin Estado" for item in estados_data]
    estados_valores = [item['total'] for item in estados_data]

    sistemas_data = tickets_filtrados.values('sistema__nombre').annotate(total=Count('id')).order_by('-total')
    sistemas_labels = [item['sistema__nombre'] or "General" for item in sistemas_data]
    sistemas_valores = [item['total'] for item in sistemas_data]

    prioridades_data = tickets_filtrados.values('prioridad__nombre').annotate(total=Count('id')).order_by('-total')
    prioridades_labels = [item['prioridad__nombre'] or "Normal" for item in prioridades_data]
    prioridades_valores = [item['total'] for item in prioridades_data]

    carga_data = tickets_filtrados.filter(~Q(estado__es_estado_cierre=True)).values('usuario_asignado__nombre_completo').annotate(total=Count('id')).order_by('-total')[:5]
    carga_labels = [item['usuario_asignado__nombre_completo'] or "Sin Asignar" for item in carga_data]
    carga_valores = [item['total'] for item in carga_data]

    sla_agentes_ordenados = sorted(sla_agentes_dict.items(), key=lambda item: item[1]['total_cerrados'], reverse=True)
    sla_agentes_labels = [item[0] for item in sla_agentes_ordenados]
    sla_agentes_valores = [int((item[1]['cumplidos'] / item[1]['total_cerrados']) * 100) if item[1]['total_cerrados'] > 0 else 100 for item in sla_agentes_ordenados]

    sla_sistemas_labels = list(sla_sistemas_dict.keys())
    sla_sistemas_valores = [int((sla_sistemas_dict[sist]['cumplidos'] / sla_sistemas_dict[sist]['total_cerrados']) * 100) for sist in sla_sistemas_labels]

    tiempo_labels = list(sla_sistemas_dict.keys())
    tiempo_valores = [int(sla_sistemas_dict[sist]['total_minutos'] / sla_sistemas_dict[sist]['total_cerrados']) if sla_sistemas_dict[sist]['total_cerrados'] > 0 else 0 for sist in tiempo_labels]

    impacto_data = tickets_filtrados.values('sistema__nombre', 'impacto_proceso').annotate(total=Count('id')).order_by('sistema__nombre')
    sistemas_unicos = list(set([item['sistema__nombre'] or "General" for item in impacto_data]))
    impacto_caido = {sist: 0 for sist in sistemas_unicos}
    impacto_parcial = {sist: 0 for sist in sistemas_unicos}
    impacto_funcional = {sist: 0 for sist in sistemas_unicos}
    
    for item in impacto_data:
        sist = item['sistema__nombre'] or "General"
        imp = str(item['impacto_proceso']).strip().upper()
        if "TOTALMENTE" in imp or "CAÍDO" in imp or "CAIDO" in imp: impacto_caido[sist] += item['total']
        elif "PARCIALMENTE" in imp or "PARCIAL" in imp: impacto_parcial[sist] += item['total']
        else: impacto_funcional[sist] += item['total']

    context = {
        'total_tickets': total_tickets, 'pendientes': pendientes, 'resueltos': resueltos, 'sla_porcentaje': sla_porcentaje,
        'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'), 'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
        'tendencias_labels': tendencias_labels, 'tendencias_valores': tendencias_valores,
        'estados_labels': estados_labels, 'estados_valores': estados_valores,
        'sistemas_labels': sistemas_labels, 'sistemas_valores': sistemas_valores,
        'prioridades_labels': prioridades_labels, 'prioridades_valores': prioridades_valores,
        'carga_labels': carga_labels, 'carga_valores': carga_valores,
        'tiempo_labels': tiempo_labels, 'tiempo_valores': tiempo_valores,
        'sla_agentes_labels': sla_agentes_labels, 'sla_agentes_valores': sla_agentes_valores,
        'sla_sistemas_labels': sla_sistemas_labels, 'sla_sistemas_valores': sla_sistemas_valores,
        'impacto_labels': sistemas_unicos,
        'impacto_caido': [impacto_caido[sist] for sist in sistemas_unicos],
        'impacto_parcial': [impacto_parcial[sist] for sist in sistemas_unicos],
        'impacto_funcional': [impacto_funcional[sist] for sist in sistemas_unicos],
    }
    if request.headers.get('HX-Request'): return render(request, 'tickets/dashboard_partials.html', context)
    return render(request, 'tickets/panel_dashboard.html', context)

@login_required
def panel_ticket_create(request):
    if request.method == "POST":
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        sistema_id = request.POST.get("sistema")
        modulo_id = request.POST.get("modulo")
        categoria_id = request.POST.get("categoria")
        prioridad_id = request.POST.get("prioridad")
        codigo_error = request.POST.get("codigo_error")
        medio_ingreso = request.POST.get("medio_ingreso", "portal")
        primer_estado = Estado.objects.order_by('orden').first()
        impacto_val = request.POST.get("impacto")
        
        nuevo_ticket = Ticket.objects.create(
            titulo=titulo, descripcion=descripcion, sistema_id=sistema_id if sistema_id else None,
            modulo_id=modulo_id if modulo_id else None, categoria_id=categoria_id if categoria_id else None,
            prioridad_id=prioridad_id if prioridad_id else None, estado=primer_estado, codigo_error=codigo_error,
            medio_ingreso=medio_ingreso, impacto_proceso=impacto_val, usuario_reporta=request.user
        )
        return redirect('panel_ticket_detail', pk=nuevo_ticket.id)

    context = {
        'sistemas': Sistema.objects.filter(activo=True),
        'categorias': Categoria.objects.all().order_by('nombre'),
        'prioridades': Prioridad.objects.all(),
    }
    return render(request, 'tickets/create.html', context)

@login_required
def panel_ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    action = request.GET.get('action', '')

    if request.method == "GET":
        if action == "edit_info":
            return render(request, 'tickets/partials/edit_form.html', {'ticket': ticket, 'sistemas': Sistema.objects.filter(activo=True)})
        elif action == "view_info":
            return render(request, 'tickets/partials/view_info.html', {'ticket': ticket})
        
        context = {
            'ticket': ticket, 'estados': Estado.objects.all().order_by('orden'), 'prioridades': Prioridad.objects.all(),
            'tecnicos': Usuario.objects.filter(rol='tecnico') or Usuario.objects.filter(is_staff=True) or Usuario.objects.all(),
        }
        return render(request, 'tickets/detail.html', context)

    if request.method == "POST":
        if action == "update_info":
            ticket.titulo = request.POST.get("titulo")
            ticket.descripcion = request.POST.get("descripcion")
            sistema_id = request.POST.get("sistema")
            ticket.sistema_id = sistema_id if sistema_id else None
            ticket.save()
            return render(request, 'tickets/partials/view_info.html', {'ticket': ticket})

        old_est = ticket.estado
        old_archivado = ticket.archivado
        estado_id = request.POST.get("estado")
        prioridad_id = request.POST.get("prioridad")
        causa_raiz = request.POST.get("causa_raiz")
        solucion_aplicada = request.POST.get("solucion_aplicada")

        if "correos_seguimiento" in request.POST:
            ticket.correos_seguimiento = request.POST.get("correos_seguimiento", "").strip()
        if estado_id: ticket.estado_id = estado_id
        if prioridad_id: ticket.prioridad_id = prioridad_id
        if "usuario_asignado" in request.POST:
            usuario_asignado_id = request.POST.get("usuario_asignado")
            ticket.usuario_asignado_id = usuario_asignado_id if usuario_asignado_id else None
        if causa_raiz is not None: ticket.causa_raiz = causa_raiz
        if solucion_aplicada is not None: ticket.solucion_aplicada = solucion_aplicada
        if "archivado" in request.POST or request.headers.get('HX-Request'):
            ticket.archivado = request.POST.get("archivado") == "true"
        
        ticket.save()

        if old_archivado != ticket.archivado:
            accion_txt = "archivado e inactivado" if ticket.archivado else "restaurado y devuelto a la lista activa"
            ChatterEntry.objects.create(ticket=ticket, tipo='sistema', autor=request.user, contenido=f"📁 El ticket ha sido {accion_txt} por el operador.")

        if old_est != ticket.estado or "usuario_asignado" in request.POST:
            _handle_state_change(ticket, old_est, ticket.estado, request.user)

        if "estado" in request.POST or "usuario_asignado" in request.POST or "prioridad" in request.POST:
            notas = ChatterEntry.objects.filter(ticket=ticket).order_by('-fecha_creacion')
            return render(request, 'tickets/partials/chatter.html', {'notas': disabled_notas if 'disabled_notas' in locals() else context_notas if 'context_notas' in locals() else notas})
            
        return HttpResponse(status=204)

@login_required
def panel_conocimiento_lista(request):
    query = request.GET.get('q', '').strip()
    soluciones = ConocimientoEntry.objects.select_related('sistema', 'ticket_origen').all()
    if query:
        soluciones = soluciones.filter(Q(titulo__icontains=query) | Q(descripcion_problema__icontains=query) | Q(codigo_error__icontains=query) | Q(solucion_aplicada__icontains=query))
    if request.headers.get('HX-Request'): return render(request, 'conocimiento/partials/soluciones_loop.html', {'soluciones': soluciones})
    return render(request, 'conocimiento/lista.html', {'soluciones': soluciones})

@login_required
def panel_conocimiento_crear(request):
    if request.method == "POST":
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion_problema")
        solucion_txt = request.POST.get("solucion_aplicada")
        causa = request.POST.get("causa_raiz")
        codigo = request.POST.get("codigo_error")
        sistema_id = request.POST.get("sistema")
        video_url = request.POST.get("video_url")
        documento_url = request.POST.get("documento_url")
        palabras_clave = request.POST.get("palabras_clave")
        
        if titulo and descripcion and solucion_txt:
            ConocimientoEntry.objects.create(
                titulo=titulo, descripcion_problema=descripcion, solucion_aplicada=solucion_txt, causa_raiz=causa,
                codigo_error=codigo, sistema_id=sistema_id if sistema_id else None, video_url=video_url if video_url else None,
                documento_url=documento_url if documento_url else None, palabras_clave=palabras_clave if palabras_clave else None
            )
        return HttpResponse('<script>window.location.reload();</script>')
    return render(request, 'conocimiento/partials/modal_crear.html', {'sistemas': Sistema.objects.filter(activo=True)})

# ─────────────────────────────────────────────
#  CMDB - SISTEMAS (MESA DE GOBIERNO CMDB)
# ─────────────────────────────────────────────

Aquí tienes la función panel_config_sistemas completa, reestructurada desde cero con una lógica de retornos simplificada y robusta.

El problema de que no paginara (y que no arrojara errores en consola ni en logs) se debe a que la petición de la página 2 incluye los parámetros de búsqueda gracias al hx-include. Al simplificar los if condicionales del final, garantizamos que cualquier petición que requiera datos dinámicos (ya sea scroll, filtro o buscador) devuelva exactamente los renglones parciales sin colisionar.

🛠️ Código Completo para views.py
Reemplaza tu función actual por esta versión limpia:

Python
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .models import Sistema, Usuario  # Asegúrate de que los nombres de importación sean correctos

@login_required
def panel_config_sistemas(request):
    """
    🖥️ CATÁLOGO DE SISTEMAS: Lista y crea de forma asíncrona con Gobierno Técnico Completo.
    Soporta búsqueda multivariable cruzada y scroll infinito estable vía HTMX.
    """
    if request.user.rol != 'admin': 
        return HttpResponse("No autorizado", status=403)

    # =========================================================================
    # 📥 1. PROCESAMIENTO DE CREACIÓN (POST TRADICIONAL / FORMULARIO)
    # =========================================================================
    if request.method == "POST" and not request.headers.get('HX-Request'):
        nombre = request.POST.get("nombre")
        activo = True if request.POST.get("activo") else False
        objetivo_descripcion = request.POST.get('objetivo_descripcion')
        acceso_recurso = request.POST.get('acceso_recurso')
        servidor_alojamiento = request.POST.get('servidor_alojamiento')
        informacion_tecnica = request.POST.get('informacion_tecnica')
        version = request.POST.get('version')
        documentacion = request.POST.get('documentacion')
        nombre_bd = request.POST.get('nombre_bd')
        formato_sistema = request.POST.get('formato_sistema')
        ubicacion_servidor = request.POST.get('ubicacion_servidor')
        plazo_conservacion = request.POST.get('plazo_conservacion')
        fecha_respaldo = request.POST.get('fecha_respaldo')
        formato_respaldo = request.POST.get('formato_respaldo')
        medio_respaldo = request.POST.get('medio_respaldo')
        observaciones = request.POST.get('observaciones')
        
        cifra_usuarios_raw = request.POST.get('cifra_usuarios')
        cifra_usuarios = int(cifra_usuarios_raw) if cifra_usuarios_raw and cifra_usuarios_raw.isdigit() else None

        desarrollador_id = request.POST.get('desarrollado_por')
        desarrollador_by = Usuario.objects.filter(id=desarrollador_id).first() if desarrollador_id else None
        resguardo_id = request.POST.get('responsable_resguardo')
        responsable_res = Usuario.objects.filter(id=resguardo_id).first() if resguardo_id else None

        if nombre:
            Sistema.objects.create(
                nombre=nombre, activo=activo, objetivo_descripcion=objetivo_descripcion, acceso_recurso=acceso_recurso,
                servidor_alojamiento=servidor_alojamiento, informacion_tecnica=informacion_tecnica, version=version,
                cifra_usuarios=cifra_usuarios, documentacion=documentacion, nombre_bd=nombre_bd, formato_sistema=formato_sistema,
                ubicacion_servidor=ubicacion_servidor, plazo_conservacion=plazo_conservacion, fecha_respaldo=fecha_respaldo,
                formato_respaldo=formato_respaldo, medio_respaldo=medio_respaldo, observaciones=observaciones,
                desarrollador_by=desarrollador_by, responsable_resguardo=responsable_res
            )

    # =========================================================================
    # 🔍 2. LÓGICA DE BÚSQUEDA Y CONTROL DE ESTADOS (GET / HTMX)
    # =========================================================================
    query = request.GET.get('q_sistema', '').strip()
    
    # Control estricto del Checkbox para evitar desfaces en la paginación (Página 2+)
    if request.headers.get('HX-Request') and 'page' in request.GET:
        solo_activos = request.GET.get('solo_activos') == 'true'
    else:
        solo_activos = request.GET.get('solo_activos', 'true') == 'true'

    # Optimización del QuerySet base
    sistemas_list = Sistema.objects.select_related('desarrollado_por', 'responsable_resguardo').all()

    # Filtro 1: Mostrar sólo activos
    if solo_activos:
        sistemas_list = sistemas_list.filter(activo=True)

    # Filtro 2: Búsqueda cruzada inteligente
    if query:
        sistemas_list = sistemas_list.filter(
            Q(nombre__icontains=query) |
            Q(objetivo_descripcion__icontains=query) |
            Q(servidor_alojamiento__icontains=query) |
            Q(informacion_tecnica__icontains=query) |
            Q(desarrollado_por__nombre_completo__icontains=query) |
            Q(responsable_resguardo__nombre_completo__icontains=query) |
            Q(desarrollado_por__region_zona__icontains=query)
        )

    # Garantizar orden estable para evitar saltos o duplicados en el Paginador
    sistemas_list = sistemas_list.order_by('nombre')

    # =========================================================================
    # 📦 3. CONFIGURACIÓN DEL PAGINADOR (SCROLL INFINITO)
    # =========================================================================
    paginator = Paginator(sistemas_list, 15)  # Bloques de 15 elementos
    page_number = request.GET.get('page', 1)
    
    try:
        sistemas = paginator.page(page_number)
    except PageNotAnInteger:
        sistemas = paginator.page(1)
    except EmptyPage:
        # Si HTMX pide una página que excede el total, devolvemos respuesta vacía para detener el trigger 'revealed'
        if request.headers.get('HX-Request'): 
            return HttpResponse("")
        sistemas = paginator.page(paginator.num_pages)
    
    # Catálogo de técnicos para el modal de creación rápida
    try:
        tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'admin']).order_by('nombre_completo')
    except Exception:
        tecnicos = Usuario.objects.all().order_by('id')
    
    context = {
        'sistemas': sistemas, 
        'tecnicos': tecnicos,
        'q_sistema': query,
        'solo_activos': solo_activos
    }

    # =========================================================================
    # 🎯 4. CONTROL DE RESPUESTAS SEGÚN LA PETICIÓN (HTMX vs TRADICIONAL)
    # =========================================================================
    if request.headers.get('HX-Request'):
        # ACCIÓN A: Si la petición trae 'page', 'q_sistema' o 'solo_activos', 
        # devolvemos UNICAMENTE los renglones correspondientes.
        if 'page' in request.GET or 'q_sistema' in request.GET or 'solo_activos' in request.GET:
            return render(request, 'configuracion/partials/sistemas_rows.html', context)
        
        # ACCIÓN B: Carga limpia del módulo completo desde el panel dinámico lateral
        return render(request, 'configuracion/partials/sistemas.html', context)
        
    # ACCIÓN C: Carga tradicional directa por barra de navegación
    return render(request, 'configuracion/panel.html', context)

    
    

@login_required
def panel_config_sistema_crear_modal(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    try:
        tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'admin']).order_by('nombre_completo')
    except Exception:
        tecnicos = Usuario.objects.filter(is_staff=True).order_by('username')
    return render(request, 'configuracion/partials/modal_sistema_crear.html', {'tecnicos': tecnicos})

@login_required
def panel_config_sistema_editar_modal(request, pk):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    sistema = get_object_or_404(Sistema, pk=pk)
    try:
        tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'admin']).order_by('nombre_completo')
    except Exception:
        tecnicos = Usuario.objects.all().order_by('id')
    return render(request, 'configuracion/partials/modal_sistema_editar.html', {'sistema': sistema, 'tecnicos': tecnicos})

@login_required
def panel_config_sistema_actualizar(request, pk):
    """
    🔄 ACCIÓN HTMX: Procesa los cambios enviados desde el modal de edición.
    """
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
        
    sistema = get_object_or_404(Sistema, pk=pk)

    if request.method == "POST":
        sistema.nombre = request.POST.get("nombre")
        sistema.activo = True if request.POST.get("activo") else False
        sistema.version = request.POST.get('version')
        sistema.formato_sistema = request.POST.get('formato_sistema')
        
        # 🚀 CORRECCIÓN AQUÍ: Cambiamos 'objective_description' por 'objetivo_descripcion'
        sistema.objetivo_descripcion = request.POST.get('objetivo_descripcion')
        
        cifra_raw = request.POST.get('cifra_usuarios')
        sistema.cifra_usuarios = int(cifra_raw) if cifra_raw and cifra_raw.isdigit() else None
        
        sistema.acceso_recurso = request.POST.get('acceso_recurso')
        sistema.servidor_alojamiento = request.POST.get('servidor_alojamiento')
        sistema.ubicacion_servidor = request.POST.get('ubicacion_servidor')
        sistema.nombre_bd = request.POST.get('nombre_bd')
        sistema.informacion_tecnica = request.POST.get('informacion_tecnica')
        sistema.documentacion = request.POST.get('documentacion')
        
        desarrollador_id = request.POST.get('desarrollado_por')
        sistema.desarrollado_por = Usuario.objects.filter(id=desarrollador_id).first() if desarrollador_id else None

        resguardo_id = request.POST.get('responsable_resguardo')
        sistema.responsable_resguardo = Usuario.objects.filter(id=resguardo_id).first() if resguardo_id else None
        
        sistema.fecha_respaldo = request.POST.get('fecha_respaldo')
        sistema.formato_respaldo = request.POST.get('formato_respaldo')
        sistema.medio_respaldo = request.POST.get('medio_respaldo')
        sistema.plazo_conservacion = request.POST.get('plazo_conservacion')
        sistema.observaciones = request.POST.get('observaciones')
        
        sistema.save()

    sistemas = Sistema.objects.all().order_by('nombre')
    try:
        tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'admin']).order_by('nombre_completo')
    except Exception:
        tecnicos = Usuario.objects.all().order_by('id')

    return render(request, 'configuracion/partials/sistemas.html', {'sistemas': sistemas, 'tecnicos': tecnicos})



@login_required
@require_http_methods(["POST"])
def panel_config_sistema_eliminar(request, pk):
    sistema = get_object_or_404(Sistema, pk=pk)
    if Ticket.objects.filter(sistema=sistema).exists() or Modulo.objects.filter(sistemas=sistema).exists():
        return HttpResponse('<script>alert("❌ No se puede eliminar: Este sistema tiene módulos o tickets vinculados.");</script>', status=200)
    sistema.delete()
    sistemas = Sistema.objects.all().order_by('nombre')
    return render(request, 'configuracion/partials/sistemas.html', {'sistemas': sistemas})


# ─────────────────────────────────────────────
#  CMDB - MÓDULOS (RELACIÓN COMPARTIDA MANY-TO-MANY)
# ─────────────────────────────────────────────

@login_required
def panel_config_modulos(request):
    """
    📦 CATÁLOGO DE MÓDULOS: Maneja listado y asignaciones de múltiples sistemas padres.
    """
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        sistema_id = request.POST.get("sistema") 

        if nombre:
            modulo = Modulo.objects.create(nombre=nombre, descripcion=descripcion, activo=True)
            if sistema_id:
                sistema = get_object_or_404(Sistema, id=sistema_id)
                sistema.modulos.add(modulo) 

    modulos = Modulo.objects.all().order_by('nombre')
    sistemas_list = Sistema.objects.filter(activo=True).order_by('nombre')
    
    return render(request, 'configuracion/partials/modulos.html', {
        'modulos': modulos,
        'sistemas_list': sistemas_list
    })

@login_required
@require_http_methods(["POST"])
def panel_config_modulo_toggle_activo(request, pk):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    modulo = get_object_or_404(Modulo, pk=pk)
    modulo.activo = not modulo.activo
    modulo.save()
    
    modulos = Modulo.objects.all().order_by('nombre')
    sistemas_list = Sistema.objects.filter(activo=True).order_by('nombre')
    
    return render(request, 'configuracion/partials/modulos.html', {
        'modulos': modulos,
        'sistemas_list': sistemas_list
    })

@login_required
def ajax_cargar_modulos(request):
    """🔍 Retorna los módulos compartidos asociados al sistema seleccionado """
    sistema_id = request.GET.get('sistema')
    if not sistema_id: return HttpResponse('<option value="">— Selecciona Módulo —</option>')
    
    sistema = get_object_or_404(Sistema, id=sistema_id)
    modulos = sistema.modulos.filter(activo=True).order_by('nombre')
    
    if modulos.exists():
        options = '<option value="">— Selecciona Módulo —</option>'
        for mod in modulos:
            options += f'<option value="{mod.id}">{mod.nombre}</option>'
    else:
        options = '<option value="">No aplica (Sistema sin submódulos)</option>'
        
    return HttpResponse(options)

# ─────────────────────────────────────────────
#  CATEGORÍAS Y USUARIOS
# ─────────────────────────────────────────────

@login_required
def panel_config_categorias(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        if nombre: Categoria.objects.create(nombre=nombre)
    categorias = Categoria.objects.all().order_by('nombre')
    return render(request, 'configuracion/partials/categorias.html', {'categorias': categorias})

@login_required
@require_http_methods(["POST"])
def panel_config_categoria_eliminar(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if Ticket.objects.filter(categoria=categoria).exists():
        return HttpResponse('<script>alert("❌ No se puede eliminar: Esta categoría tiene tickets activos vinculados.");</script>', status=200)
    categoria.delete()
    categorias = Categoria.objects.all().order_by('nombre')
    return render(request, 'configuracion/partials/categorias.html', {'categorias': categorias})

@login_required
@require_http_methods(["POST"])
def panel_ticket_add_comentario(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    contenido = request.POST.get("contenido", "").strip()
    
    if contenido:
        nueva_nota = ChatterEntry.objects.create(ticket=ticket, tipo='comentario', autor=request.user, contenido=contenido)
        lista_correos = []
        if ticket.usuario_reporta and getattr(ticket.usuario_reporta, 'correo_electronico', None):
            lista_correos.append(ticket.usuario_reporta.correo_electronico)
        if ticket.usuario_asignado and getattr(ticket.usuario_asignado, 'correo_electronico', None):
            lista_correos.append(ticket.usuario_asignado.correo_electronico)
        if ticket.correos_seguimiento:
            adicionales = [c.strip() for c in ticket.correos_seguimiento.split(',') if c.strip()]
            lista_correos.extend(adicionales)

        lista_correos = list(set(lista_correos))
        if lista_correos:
            folio_ticket = getattr(ticket, 'folio', ticket.id)
            titulo_ticket = getattr(ticket, 'titulo', 'Soporte Técnico')
            nombre_remitente = getattr(request.user, 'nombre_completo', request.user.username)
            asunto = f"🔔 Actualización en Ticket #{folio_ticket} - {titulo_ticket}"
            mensaje_texto = f"El usuario {nombre_remitente} ha agregado una nueva nota:\n\n\"{contenido}\""
            
            try:
                send_mail(subject=asunto, message=mensaje_texto, from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=lista_correos, fail_silently=True)
            except Exception: pass

        return render(request, 'tickets/partials/chatter_loop.html', {'notas': [nueva_nota]})
    return HttpResponse(status=400)

@login_required
def panel_usuarios_list(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    query = request.GET.get('q', '').strip()
    usuarios = Usuario.objects.all().order_by('nombre_completo')
    
    if query:
        usuarios = usuarios.filter(Q(nombre_completo__icontains=query) | Q(correo_electronico__icontains=query) | Q(numero_empleado__icontains=query) | Q(puesto_cargo__icontains=query) | Q(cct__icontains=query))

    context = {'usuarios': usuarios}
    if request.headers.get('HX-Request'): return render(request, 'usuarios/partials/usuarios_render_search.html', context)
    return render(request, 'usuarios/lista.html', context)

@login_required
@require_http_methods(["POST"])
def panel_usuario_cambiar_rol(request, user_id):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    usuario = get_object_or_404(Usuario, pk=user_id)
    nuevo_rol = request.POST.get("rol")
    if nuevo_rol in ['admin', 'tecnico', 'usuario']:
        usuario.rol = nuevo_rol
        usuario.is_staff = True if nuevo_rol == 'admin' else False
        usuario.save()
        return HttpResponse(status=200)
    return HttpResponse(status=400)

@login_required
@require_http_methods(["POST"])
def panel_usuario_toggle_activo(request, user_id):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    usuario = get_object_or_404(Usuario, pk=user_id)
    usuario.activo = not usuario.activo
    usuario.save()

    if usuario.activo:
        return HttpResponse(f'<button hx-post="/api/panel/usuarios/{usuario.id}/toggle/" hx-headers=\'{{"X-CSRFToken": "{request.META.get("CSRF_COOKIE")}"}}\' hx-target="this" hx-swap="outerHTML" class="px-2.5 py-1 text-[10px] font-bold rounded-full border bg-emerald-50 text-emerald-700 border-emerald-200 hover:bg-emerald-100">● Activo</button>')
    return HttpResponse(f'<button hx-post="/api/panel/usuarios/{usuario.id}/toggle/" hx-headers=\'{{"X-CSRFToken": "{request.META.get("CSRF_COOKIE")}"}}\' hx-target="this" hx-swap="outerHTML" class="px-2.5 py-1 text-[10px] font-bold rounded-full border bg-slate-50 text-slate-500 border-slate-200 hover:bg-slate-100">○ Inactivo</button>')

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
# 🚀 Asegúrate de importar el modelo Departamento si no está arriba:
from .models import Usuario, Departamento 

@login_required
def panel_usuario_editar(request, user_id):
    if request.user.rol != 'admin': 
        return HttpResponse("No autorizado", status=403)
    
    usuario = get_object_or_404(Usuario, pk=user_id)
    
    if request.method == "POST":
        # ... (Tus asignaciones actuales: nombre_completo, email, extension, etc.)
        usuario.nombre_completo = request.POST.get("nombre_completo")
        usuario.correo_electronico = request.POST.get("email")
        usuario.extension = request.POST.get("extension")
        usuario.puesto_cargo = request.POST.get("puesto_cargo")
        usuario.numero_empleado = request.POST.get("numero_empleado")
        usuario.cct = request.POST.get("cct")
        usuario.region_zona = request.POST.get("region_zona")
        usuario.nivel_educativo = request.POST.get("nivel_educativo")
        
        # 🏢 1. PROCESAR EL DEPARTAMENTO SELECCIONADO EN EL POST:
        dept_id = request.POST.get("departamento")
        if dept_id:
            usuario.departamento_id = dept_id  # Guardamos la relación usando el ID directo
        else:
            usuario.departamento = None  # Si eligió "Sin Asignar"
            
        estado_raw = request.POST.get("estado") == "True"
        usuario.activo = estado_raw
        usuario.rol = request.POST.get("rol")
        usuario.save()
        
        # Devolvemos la fila actualizada para HTMX
        return render(request, 'usuarios/partials/usuarios_row.html', {
            'u': usuario,
            'usuario': usuario
        })
        
    # 🏢 2. ENVIAR LOS DEPARTAMENTOS DISPONIBLES AL MODAL (GET):
    # Traemos los que están activos O el que ya tenga asignado el usuario (aunque esté archivado)
    departamentos = Departamento.objects.filter(
        Q(activo=True) | Q(id=usuario.departamento_id)
    ).distinct().order_by('nombre')
    
    return render(request, 'usuarios/partials/modal_editar.html', {
        'usuario': usuario,
        'departamentos': departamentos  # 🚀 CRUCIAL para que el {% for %} funcione
    })


        

@login_required
def panel_usuario_importar_csv(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    if request.method == "POST":
        csv_file = request.FILES.get('file')
        if not csv_file or not csv_file.name.endswith('.csv'): return HttpResponse("Formato inválido. Sube un archivo .csv", status=400)

        try:
            data_set = csv_file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            csv_file.seek(0)
            data_set = csv_file.read().decode('latin-1')

        io_string = io.StringIO(data_set)
        next(io_string) 

        for row in csv.reader(io_string, delimiter=','):
            if not row or len(row) < 2: continue
            correo = row[0].strip()
            nombre = row[1].strip()
            while len(row) < 9: row.append("")
                
            num_emp = row[2].strip() if row[2].strip() else None
            puesto = row[3].strip() if row[3].strip() else None
            cct_val = row[4].strip() if row[4].strip() else None
            region = row[5].strip() if row[5].strip() else None
            nivel = row[6].strip() if row[6].strip() else None
            tel_val = row[7].strip() if row[7].strip() else None
            ext_val = row[8].strip() if row[8].strip() else None

            if correo and nombre:
                try:
                    usuario_obj, creado = Usuario.objects.get_or_create(
                        correo_electronico=correo,
                        defaults={'nombre_completo': nombre, 'numero_empleado': num_emp, 'puesto_cargo': puesto, 'cct': cct_val, 'region_zona': region, 'nivel_educativo': nivel, 'telefono': tel_val, 'extension': ext_val, 'rol': 'usuario'}
                    )
                    if creado:
                        usuario_obj.set_password(num_emp if num_emp else "Seech2026*")
                        usuario_obj.save()
                    else:
                        usuario_obj.numero_empleado = num_emp
                        usuario_obj.puesto_cargo = puesto
                        usuario_obj.cct = cct_val
                        usuario_obj.region_zona = region
                        usuario_obj.nivel_educativo = nivel
                        usuario_obj.telefono = tel_val
                        usuario_obj.extension = ext_val
                        usuario_obj.save()
                except Exception as row_error:
                    continue

        return render(request, 'usuarios/partials/usuarios_row.html', {'usuarios': Usuario.objects.all().order_by('-id'), 'usuario': request.user})
    return render(request, 'usuarios/partials/modal_csv.html')

@login_required
def panel_reportes_avanzados(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado_id = request.GET.get('estado')
    asignado_id = request.GET.get('asignado')
    categoria_id = request.GET.get('categoria')
    sistema_id = request.GET.get('sistema')
    modulo_id = request.GET.get('modulo')
    impacto = request.GET.get('impacto')
    region = request.GET.get('region')
    archivado_filtro = request.GET.get('archivado', 'false')

    qs = Ticket.objects.select_related('sistema', 'modulo', 'prioridad', 'estado', 'categoria', 'usuario_asignado').all()

    if fecha_inicio: qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin: qs = qs.filter(fecha_creacion__date__lte=fecha_fin)
    if estado_id: qs = qs.filter(estado_id=estado_id)
    if asignado_id: qs = qs.filter(usuario_asignado_id=asignado_id)
    if categoria_id: qs = qs.filter(categoria_id=categoria_id)
    if sistema_id: qs = qs.filter(sistema_id=sistema_id)
    if modulo_id: qs = qs.filter(modulo_id=modulo_id)
    if impacto: qs = qs.filter(impacto_proceso=impacto)
    if region: qs = qs.filter(usuario_reporta__region_zona__icontains=region)
    if archivado_filtro == 'true': qs = qs.filter(archivado=True)
    elif archivado_filtro == 'false': qs = qs.filter(archivado=False)

    sistemas_data = qs.values('sistema__nombre').annotate(total=Count('id')).order_by('-total')
    sistemas_labels = [item['sistema__nombre'] or 'Sin Sistema' for item in sistemas_data]
    sistemas_valores = [item['total'] for item in sistemas_data]

    estados_data = qs.values('estado__nombre').annotate(total=Count('id')).order_by('-total')
    estados_labels = [item['estado__nombre'] or 'Sin Estado' for item in estados_data]
    estados_valores = [item['total'] for item in estados_data]

    context = {
        'tickets': qs.order_by('-fecha_creacion')[:200], 'total_filtrado': qs.count(),
        'estados': Estado.objects.all().order_by('orden'), 'categorias': Categoria.objects.all().order_by('nombre'),
        'sistemas': Sistema.objects.filter(activo=True), 'tecnicos': Usuario.objects.filter(rol='tecnico'),
        'sistemas_labels': json.dumps(sistemas_labels), 'sistemas_valores': json.dumps(sistemas_valores),
        'estados_labels': json.dumps(estados_labels), 'estados_valores': json.dumps(estados_valores),
        'current_archivado': archivado_filtro,
    }
    if request.headers.get('HX-Request'): return render(request, 'configuracion/partials/reportes_resultados.html', context)
    return render(request, 'configuracion/reportes.html', context)

@login_required
def exportar_reporte_csv(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado_id = request.GET.get('estado')
    asignado_id = request.GET.get('asignado')
    categoria_id = request.GET.get('categoria')
    sistema_id = request.GET.get('sistema')
    modulo_id = request.GET.get('modulo')
    impacto = request.GET.get('impacto')
    region = request.GET.get('region')
    archivado_filtro = request.GET.get('archivado', 'false')

    qs = Ticket.objects.select_related('sistema', 'modulo', 'estado', 'categoria', 'usuario_asignado', 'usuario_reporta').all().order_by('-fecha_creacion')

    if fecha_inicio: qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin: qs = qs.filter(fecha_creacion__date__lte=fecha_fin)
    if estado_id: qs = qs.filter(estado_id=estado_id)
    if asignado_id: qs = qs.filter(usuario_asignado_id=asignado_id)
    if categoria_id: qs = qs.filter(categoria_id=categoria_id)
    if sistema_id: qs = qs.filter(sistema_id=sistema_id)
    if modulo_id: qs = qs.filter(modulo_id=modulo_id)
    if impacto: qs = qs.filter(impacto_proceso=impacto)
    if region: qs = qs.filter(usuario_reporta__region_zona__icontains=region)
    if archivado_filtro == 'true': qs = qs.filter(archivado=True)
    elif archivado_filtro == 'false': qs = qs.filter(archivado=False)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="Reporte_BI_Avanzado_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Folio', 'Título', 'Sistema', 'Módulo', 'Categoría', 'Estado', 'Asignado A', 'Fecha Creación', 'Fecha Asignación', 'Fecha 1ra Respuesta', 'Fecha Resolución', 'Fecha Cierre', 'Tiempo Atención (Min)', 'Tiempo Pausa (Min)'])
    
    for tk in qs:
        def _fmt_dt(dt): return dt.strftime('%d/%m/%Y %H:%M') if dt else '—'
        writer.writerow([tk.folio, tk.titulo, tk.sistema.nombre if tk.sistema else '—', tk.modulo.nombre if tk.modulo else '—', tk.categoria.nombre if tk.categoria else '—', tk.estado.nombre if tk.estado else '—', tk.usuario_asignado.nombre_completo if tk.usuario_asignado else 'Sin Asignar', _fmt_dt(tk.fecha_creacion), _fmt_dt(tk.fecha_asignacion), _fmt_dt(tk.fecha_primera_respuesta), _fmt_dt(tk.fecha_resolucion), _fmt_dt(tk.fecha_cierre), tk.tiempo_atencion_minutos or 0, tk.tiempo_pausa_minutos or 0])
    return response

@login_required
def panel_conocimiento_detalle(request, pk):
    solucion = get_object_or_404(ConocimientoEntry.objects.select_related('sistema', 'modulo', 'ticket_origen'), pk=pk)
    ConocimientoEntry.objects.filter(pk=pk).update(veces_consultado=F('veces_consultado') + 1)
    solucion.refresh_from_db()
    return render(request, 'conocimiento/detalle.html', {'solucion': solucion})

@login_required
def panel_conocimiento_eliminar(request, pk):
    if request.method == "POST":
        get_object_or_404(ConocimientoEntry, pk=pk).delete()
    return redirect('panel_conocimiento_lista')

@login_required
def panel_conocimiento_importar_csv(request):
    if request.method == "POST":
        csv_file = request.FILES.get('file')
        if not csv_file: return HttpResponse("No se subió ningún archivo.", status=400)
        if not csv_file.name.endswith('.csv'): return HttpResponse("Formato inválido. Sube un archivo .csv", status=400)

        encodings = ['utf-8-sig', 'latin-1', 'windows-1252', 'utf-8']
        data_set = None
        for encoding in encodings:
            try:
                csv_file.seek(0)
                data_set = csv_file.read().decode(encoding)
                break
            except UnicodeDecodeError: continue

        if data_set is None: return HttpResponse("Error de codificación.", status=400)
        try:
            io_string = io.StringIO(data_set)
            primera_linea = io_string.readline()
            delimitador = ';' if ';' in primera_linea else ','
            io_string.seek(0)
            reader = csv.reader(io_string, delimiter=delimitador)
            if any(palabra in primera_linea.lower() for palabra in ['titulo', 'descrip', 'solucion']): next(reader)

            for row in reader:
                if not row or len(row) == 0: continue
                titulo_csv = row[0].strip() if len(row) > 0 else ""
                if titulo_csv:
                    ConocimientoEntry.objects.create(
                        titulo=titulo_csv, descripcion_problema=row[1].strip() if len(row) > 1 else "",
                        solucion_aplicada=row[2].strip() if len(row) > 2 else "",
                        codigo_error=row[3].strip() if (len(row) > 3 and row[3].strip()) else None,
                        causa_raiz=row[4].strip() if (len(row) > 4 and row[4].strip()) else None, sistema=None
                    )
            return HttpResponse('<script>window.location.reload();</script>')
        except Exception as e: return HttpResponse(f"Error: {str(e)}", status=500)
    return render(request, 'conocimiento/partials/modal_csv.html')

@login_required
@require_http_methods(["POST"])
def panel_usuario_eliminar(request, user_id):
    if request.user.rol != 'admin': return HttpResponse("No autorizado.", status=403)
    usuario_a_borrar = get_object_or_404(Usuario, pk=user_id)
    if usuario_a_borrar.id == request.user.id:
        return HttpResponse('<script>alert("❌ Error: No puedes eliminar tu propia cuenta."); window.location.reload();</script>')
    usuario_a_borrar.delete()
    return redirect('panel_usuarios_list')

@login_required
def panel_usuarios_exportar_excel(request):
    if request.user.rol != 'admin': 
        return HttpResponse("No autorizado", status=403)
        
    query = request.GET.get('q', '').strip()
    
    # 🚀 OPTIMIZACIÓN: select_related para traer el departamento en una sola query
    usuarios = Usuario.objects.all().select_related('departamento').order_by('nombre_completo')
    
    if query:
        usuarios = usuarios.filter(
            Q(nombre_completo__icontains=query) | 
            Q(correo_electronico__icontains=query) | 
            Q(numero_empleado__icontains=query) | 
            Q(puesto_cargo__icontains=query) | 
            Q(cct__icontains=query) |
            Q(departamento__nombre__icontains=query) # 🔍 Permite exportar el filtro si buscaron un depto
        )

    response = HttpResponse(content_type='text/csv; charset=windows-1252')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios_seech.csv"'
    
    writer = csv.writer(response, delimiter=';')
    
    # 🏢 Cabecera actualizada con 'Departamento / Area'
    writer.writerow([
        'Nombre Completo', 'Correo Electronico', 'Numero de Empleado', 
        'Puesto / Cargo', 'Departamento / Area', 'CCT', 'Region / Zona', 
        'Nivel Educativo', 'Rol de Acceso', 'Estado'
    ])

    for u in usuarios:
        # 🏢 Extraer el nombre del departamento o dejar vacío si es NULL
        nombre_depto = u.departamento.nombre if u.departamento else '—'

        writer.writerow([
            str(u.nombre_completo).encode('windows-1252', 'replace').decode('windows-1252'),
            str(u.correo_electronico).encode('windows-1252', 'replace').decode('windows-1252'),
            u.numero_empleado or '', 
            str(u.puesto_cargo or '').encode('windows-1252', 'replace').decode('windows-1252'),
            # 🚀 Inyección del Departamento formateado de forma segura:
            str(nombre_depto).encode('windows-1252', 'replace').decode('windows-1252'),
            u.cct or '', 
            str(u.region_zona or '').encode('windows-1252', 'replace').decode('windows-1252'),
            str(u.nivel_educativo or '').encode('windows-1252', 'replace').decode('windows-1252'),
            u.get_rol_display() if hasattr(u, 'get_rol_display') else u.rol, 
            'Activo' if u.activo else 'Inactivo'
        ])
        
    return response

def _tarea_enviar_correo_async(asunto, html_contenido, remitente, destino):
    import urllib.request, urllib.error
    api_key = getattr(settings, 'EMAIL_HOST_PASSWORD', '') 
    if not api_key or api_key.startswith('smtp'): return
    
    payload = {"from": "Mesa de Ayuda SEECH <notificaciones@routripcreator.com>", "to": destino, "subject": asunto, "html": html_contenido}
    try:
        data_bytes = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request("https://api.resend.com/emails", data=data_bytes, headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json", "User-Agent": "Django-SEECH-Tickets/1.0"}, method="POST")
        with urllib.request.urlopen(req, timeout=10) as response: pass
    except Exception: pass

@login_required
@require_http_methods(["POST"])
def panel_ticket_enviar_recordatorio(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if not ticket.usuario_asignado: return HttpResponse("Sin especialista asignado.", status=400)
    if ticket.estado and ticket.estado.es_estado_cierre: return HttpResponse("Ticket cerrado.", status=400)

    asunto = f"🚨 RECORDATORIO URGENTE: Ticket Pendiente [{ticket.folio}]"
    correo_destino = ticket.usuario_asignado.correo_electronico or ticket.usuario_asignado.email
    html_contenido = f"<h3>Recordatorio de atención para {ticket.folio}</h3>"

    threading.Thread(target=_tarea_enviar_correo_async, args=(asunto, html_contenido, settings.DEFAULT_FROM_EMAIL, [correo_destino])).start()
    ChatterEntry.objects.create(ticket=ticket, tipo='sistema', autor=request.user, contenido=f"🔔 Se solicitó un recordatorio urgente.")
    return HttpResponse("✓ Enviado")

@login_required
def panel_usuario_crear(request):
    if request.user.rol != 'admin': 
        return HttpResponse("No autorizado", status=403)
        
    if request.method == "POST":
        correo = (request.POST.get("correo_electronico") or request.POST.get("email") or "").strip()
        nombre = (request.POST.get("nombre_completo") or request.POST.get("nombre") or "").strip()
        
        if not correo or not nombre: 
            return HttpResponse('⚠️ Campos obligatorios.', status=200)
        if Usuario.objects.filter(correo_electronico=correo).exists(): 
            return HttpResponse('❌ Correo duplicado.', status=200)

        # 🏢 1. CAPTURAR EL DEPARTAMENTO DEL FORMULARIO
        # Si no se seleccionó ninguno, guardamos None (NULL en DB)
        dept_id = request.POST.get("departamento")
        depto_seleccionado = dept_id if dept_id else None

        nuevo = Usuario.objects.create(
            correo_electronico=correo, 
            nombre_completo=nombre, 
            numero_empleado=request.POST.get("numero_empleado"),
            puesto_cargo=request.POST.get("puesto_cargo"), 
            cct=request.POST.get("cct"), 
            region_zona=request.POST.get("region_zona"),
            nivel_educativo=request.POST.get("nivel_educativo"), 
            rol=request.POST.get("rol", "usuario"), 
            departamento_id=depto_seleccionado,  # 🚀 ASIGNACIÓN DIRECTA POR ID
            extension=request.POST.get("extension"),  # 📞 Agregado por simetría
            telefono=request.POST.get("telefono"),    # 📱 Agregado por simetría
            activo=True
        )
        
        nuevo.set_password(nuevo.numero_empleado if nuevo.numero_empleado else "Seech2026*")
        nuevo.save()
        
        return HttpResponse('<script>window.location.reload();</script>')
        
    # 🏢 2. EN EL GET: OBTENER EL CATÁLOGO DE DEPARTAMENTOS ACTIVOS
    # Esto alimenta el bucle {% for dept in departamentos %} en el HTML del modal
    departamentos = Departamento.objects.filter(activo=True).order_by('nombre')
    
    return render(request, 'usuarios/partials/modal_crear.html', {
        'departamentos': departamentos
    })

@login_required
def panel_conocimiento_crear_desde_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if request.method == "POST":
        ConocimientoEntry.objects.create(
            titulo=request.POST.get("titulo"), descripcion_problema=request.POST.get("descripcion_problema"),
            solucion_aplicada=request.POST.get("solucion_aplicada"), causa_raiz=request.POST.get("causa_raiz"),
            codigo_error=request.POST.get("codigo_error"), sistema=ticket.sistema, modulo=ticket.modulo, ticket_origen=ticket,
            video_url=request.POST.get("video_url"), documento_url=request.POST.get("documento_url"), palabras_clave=request.POST.get("palabras_clave")
        )
        return HttpResponse('<script>window.location.reload();</script>')
    return render(request, 'conocimiento/partials/modal_convertir.html', {'ticket': ticket, 'sistemas': Sistema.objects.filter(activo=True)})

@login_required
def panel_conocimiento_editar(request, entrada_id):
    solucion = get_object_or_404(ConocimientoEntry, pk=entrada_id)
    if request.method == "POST":
        solucion.titulo = request.POST.get("titulo")
        solucion.descripcion_problema = request.POST.get("descripcion_problema")
        solucion.solucion_aplicada = request.POST.get("solucion_aplicada")
        solucion.causa_raiz = request.POST.get("causa_raiz")
        solucion.codigo_error = request.POST.get("codigo_error")
        solucion.sistema_id = request.POST.get("sistema") or None
        solucion.video_url = request.POST.get("video_url") or None
        solucion.documento_url = request.POST.get("documento_url") or None
        solucion.palabras_clave = request.POST.get("palabras_clave") or None
        solucion.save()
        return HttpResponse('<script>window.location.reload();</script>')
    return render(request, 'conocimiento/partials/modal_editar.html', {'solucion': solucion, 'sistemas': Sistema.objects.filter(activo=True)})

# ─────────────────────────────────────────────
#  TRIAGE Y MÓDULO CMDB MATRICIAL
# ─────────────────────────────────────────────

@login_required
def ajax_obtener_responsables_cmdb(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if not ticket.sistema: return HttpResponse("— Requiere asignar un Sistema primero —", status=200)
    relaciones = RelacionUsuarioSistema.objects.filter(sistema=ticket.sistema).select_related('usuario')
    correos = [r.usuario.correo_electronico for r in relaciones if r.usuario.correo_electronico]
    return HttpResponse(", ".join(correos))

@login_required
def panel_config_cmdb(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    if request.method == "POST":
        usuario_id = request.POST.get("usuario_id")
        sistema_id = request.POST.get("sistema_id")
        tipo_relacion = request.POST.get("tipo_relacion", "lider_tecnico")
        if usuario_id and sistema_id:
            RelacionUsuarioSistema.objects.get_or_create(usuario_id=usuario_id, sistema_id=sistema_id, tipo_relacion=tipo_relacion)

    context = {
        'cmdb_relaciones': RelacionUsuarioSistema.objects.select_related('usuario', 'sistema').all().order_by('sistema__nombre', 'usuario__nombre_completo'),
        'sistemas_list': Sistema.objects.filter(activo=True).order_by('nombre'),
        'usuarios_list': Usuario.objects.filter(activo=True).order_by('nombre_completo'),
    }
    if request.headers.get('HX-Request') or request.method == "POST": return render(request, 'configuracion/partials/cmdb.html', context)
    return render(request, 'configuracion/panel.html', context)

@login_required
@require_http_methods(["POST"])
def panel_config_cmdb_eliminar(request, pk):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    get_object_or_404(RelacionUsuarioSistema, pk=pk).delete()
    context = {
        'cmdb_relaciones': RelacionUsuarioSistema.objects.select_related('usuario', 'sistema').all().order_by('sistema__nombre', 'usuario__nombre_completo'),
        'sistemas_list': Sistema.objects.filter(activo=True).order_by('nombre'),
        'usuarios_list': Usuario.objects.filter(activo=True).order_by('nombre_completo'),
    }
    return render(request, 'configuracion/partials/cmdb.html', context)

@login_required
def panel_config_sistema_csv_modal(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    return render(request, 'configuracion/partials/modal_sistema_csv.html')

@login_required
def panel_config_sistema_importar_csv(request):
    if request.user.rol != 'admin': return HttpResponse("No autorizado", status=403)
    if request.method == "POST":
        csv_file = request.FILES.get('file')
        if not csv_file or not csv_file.name.endswith('.csv'): return HttpResponse("Archivo inválido.", status=400)

        try:
            try: data_set = csv_file.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                csv_file.seek(0)
                data_set = csv_file.read().decode('latin-1')

            io_string = io.StringIO(data_set)
            next(io_string) 

            for row in csv.reader(io_string, delimiter=','):
                if not row or len(row) == 0: continue
                nombre = row[0].strip()
                if not nombre: continue
                while len(row) < 17: row.append("")

                try:
                    cifra_raw = row[4].strip()
                    cifra_usuarios = int(float(cifra_raw)) if cifra_raw else None
                    email_dev = row[10].strip()
                    email_resguardo = row[11].strip()

                    Sistema.objects.update_or_create(
                        nombre=nombre,
                        defaults={
                            'activo': True, 'version': row[1].strip() or None, 'formato_sistema': row[2].strip() or None,
                            'objetivo_descripcion': row[3].strip() or None, 'cifra_usuarios': cifra_usuarios, 'acceso_recurso': row[5].strip() or None,
                            'servidor_alojamiento': row[6].strip() or None, 'ubicacion_servidor': row[7].strip() or None, 'nombre_bd': row[8].strip() or None,
                            'informacion_tecnica': row[9].strip() or None, 'desarrollado_por': Usuario.objects.filter(correo_electronico=email_dev).first() if email_dev else None,
                            'responsable_resguardo': Usuario.objects.filter(correo_electronico=email_resguardo).first() if email_resguardo else None,
                            'fecha_respaldo': row[12].strip() or None, 'formato_respaldo': row[13].strip() or None, 'medio_respaldo': row[14].strip() or None,
                            'plazo_conservacion': row[15].strip() or None, 'observaciones': row[16].strip() or None
                        }
                    )
                except Exception: continue
        except Exception as e: return HttpResponse(f"Error: {str(e)}", status=500)

    sistemas = Sistema.objects.all().order_by('nombre')
    try: tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'admin']).order_by('nombre_completo')
    except Exception: tecnicos = Usuario.objects.all().order_by('id')
    return render(request, 'configuracion/partials/sistemas.html', {'sistemas': sistemas, 'tecnicos': tecnicos})


@login_required
def panel_directorio(request):
    """
    📇 DIRECTORIO INSTITUCIONAL: Vista de solo lectura para listar usuarios.
    Soporta búsqueda en tiempo real vía HTMX y paginación infinita.
    """
    query = request.GET.get('q_directorio', '').strip()

    # 🚀 OPTIMIZACIÓN CRUCIAL: select_related('departamento') trae el nombre del depto
    # en un solo JOIN de SQL en lugar de hacer una consulta por cada fila de la tabla.
    usuarios_list = Usuario.objects.all().select_related('departamento')

    # Búsqueda multivariable (Nombre, Correo, Puesto/Cargo, Extensión, Número de Empleado y Departamento)
    if query:
        usuarios_list = usuarios_list.filter(
            Q(nombre_completo__icontains=query) |
            Q(correo_electronico__icontains=query) |
            Q(puesto_cargo__icontains=query) |
            Q(extension__icontains=query) |
            Q(numero_empleado__icontains=query) |
            Q(departamento__nombre__icontains=query)  # 🚀 NUEVO: Permite buscar por Departamento en vivo
        )

    # Ordenamos alfabéticamente por nombre
    usuarios_list = usuarios_list.order_by('nombre_completo')

    # Paginación para Scroll Infinito (15 registros por tanda)
    paginator = Paginator(usuarios_list, 15)
    page_number = request.GET.get('page', 1)

    try:
        usuarios = paginator.page(page_number)
    except PageNotAnInteger:
        usuarios = paginator.page(1)
    except EmptyPage:
        if request.headers.get('HX-Request'):
            return HttpResponse("")  # Detiene el scroll si no hay más páginas
        usuarios = paginator.page(paginator.num_pages)

    context = {
        'usuarios': usuarios,
        'q_directorio': query,
    }

    # Control de respuestas HTMX vs Carga Completa
    if request.headers.get('HX-Request'):
        # Tanto el scroll infinito como la búsqueda refrescan las filas partials
        return render(request, 'directorio/partials/directorio_rows.html', context)

    return render(request, 'directorio/directorio.html', context)


@login_required
def exportar_directorio_excel(request):
    """
    📊 EXPORTAR A EXCEL: Genera un archivo .xlsx basado en los filtros aplicados en el buscador.
    """
    query = request.GET.get('q_directorio', '').strip()
    usuarios_list = Usuario.objects.all()

    # 🚀 CORREGIDO: Usando los campos exactos de tu modelo de base de datos
    if query:
        usuarios_list = usuarios_list.filter(
            Q(nombre_completo__icontains=query) |
            Q(correo_electronico__icontains=query) |
            Q(puesto_cargo__icontains=query) |
            Q(extension__icontains=query) |
            Q(numero_empleado__icontains=query)
        )

    usuarios_list = usuarios_list.order_by('nombre_completo')

    # Crear el libro de Excel en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directorio"

    # Encabezados de la tabla
    headers = ["Clave / ID", "Nombre Completo", "Correo Electrónico", "Puesto / Cargo", "Extensión", "Rol"]
    ws.append(headers)

    # Estilo básico para los encabezados (Azul corporativo)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    # Inyectar datos reales de los usuarios filtrados
    for u in usuarios_list:
        ws.append([
            u.id,  # O u.numero_empleado si prefieres usar ese como clave visible
            u.nombre_completo,
            u.correo_electronico,
            getattr(u, 'puesto_cargo', '—'),
            getattr(u, 'extension', '—'),
            getattr(u, 'rol', '—').capitalize()
        ])

    # Auto-ajustar el ancho de las columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Preparar respuesta HTTP para la descarga
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Directorio_Personal.xlsx"'
    wb.save(response)
    return response


@login_required
def panel_departamentos(request):
    """
    🖥️ Vista principal que renderiza el contenedor y la estructura del panel de departamentos.
    """
    if request.user.rol != 'admin':
        return HttpResponse("No autorizado", status=403)
    
    # Traemos todos ordenados por nombre
    departamentos = Departamento.objects.all().order_by('nombre')
    return render(request, 'departamentos/panel.html', {'departamentos': departamentos})


@login_required
def departamento_crear(request):
    """
    ➕ Crear un nuevo departamento vía HTMX
    """
    if request.user.rol != 'admin':
        return HttpResponse("No autorizado", status=403)

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        if not nombre:
            return HttpResponse("El nombre es requerido", status=400)
        
        # Crear el registro
        dept, creado = Departamento.objects.get_or_create(nombre=nombre)
        
        # Devolvemos la fila fresca para que HTMX la inserte en la tabla
        return render(request, 'departamentos/partials/departamento_row.html', {'d': dept})
    
    # Si es GET, devolvemos el formulario limpio dentro del modal
    return render(request, 'departamentos/partials/modal_crear.html')


@login_required
def departamento_editar(request, dept_id):
    """
    📝 Editar o Archivar un departamento existente
    """
    if request.user.rol != 'admin':
        return HttpResponse("No autorizado", status=403)

    dept = get_object_or_404(Departamento, pk=dept_id)

    if request.method == "POST":
        dept.nombre = request.POST.get("nombre", "").strip()
        dept.activo = request.POST.get("activo") == "True"
        dept.save()

        # HTMX reemplaza la fila vieja por esta fila actualizada
        return render(request, 'departamentos/partials/departamento_row.html', {'d': dept})

    # Si es GET, carga los datos actuales en el modal de edición
    return render(request, 'departamentos/partials/modal_editar.html', {'dept': dept})
